from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from math_eval_framework.exporter import export_evaluation_artifact
from math_eval_framework.models import EvaluationArtifact, EvaluationRun, EvaluationSummary, MathItem


def test_exporter_writes_expected_headers(tmp_path: Path) -> None:
    item = MathItem.model_validate(
        {
            "序号": 1,
            "问题": "Test question",
            "适合年级": "研究生",
            "领域类型": "分析",
            "考察知识点": ["积分"],
            "易错点": ["符号错误"],
            "思考过程/分析": "test",
            "解题过程": ["step1"],
            "最终答案": "$3$",
        }
    )
    artifact = EvaluationArtifact(
        item=item,
        summary=EvaluationSummary(
            item_sequence=1,
            model_name="qwen3.5-plus",
            run_count=8,
            correct_count=3,
            accuracy=0.375,
            generated_at=datetime(2026, 4, 4, 1, 0, 0),
        ),
        runs=[
            EvaluationRun(
                run_index=1,
                query="{}",
                raw_response='{"最终答案":"$3$"}',
                parsed_response={"最终答案": "$3$"},
                predicted_answer="$3$",
                is_correct=True,
            )
        ],
        request_config={},
        validation_results=[{"rule": "x", "status": "PASS", "detail": "ok"}],
    )
    output = tmp_path / "result.xlsx"
    export_evaluation_artifact(artifact, output)

    workbook = load_workbook(output)
    sheet = workbook["Sheet1"]
    assert sheet.cell(1, 1).value == "序号"
    assert sheet.cell(2, 1).value == 1
    assert sheet.cell(2, 17).value == 3
