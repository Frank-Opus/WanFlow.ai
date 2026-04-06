from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from .models import EvaluationArtifact


MAX_EXPORT_RUNS = 8
RUN_HEADERS = [f"模型运行{i}" for i in range(1, MAX_EXPORT_RUNS + 1)]
HEADERS = [
    "序号",
    "问题",
    "适合年级",
    "领域类型",
    "考察知识点",
    "易错点",
    "解题过程",
    "最终答案",
    *RUN_HEADERS,
    "正确次数",
    "总运行次数",
    "命中率",
]


def export_evaluation_artifact(artifact: EvaluationArtifact, output_path: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(HEADERS)

    base_row = [
        artifact.item.sequence,
        artifact.item.question,
        artifact.item.grade_level,
        artifact.item.domain_type,
        json.dumps(artifact.item.key_points, ensure_ascii=False),
        json.dumps(artifact.item.pitfalls, ensure_ascii=False),
        json.dumps(artifact.item.solution_steps, ensure_ascii=False),
        artifact.item.final_answer,
    ]

    run_cells = []
    for run in artifact.runs[:MAX_EXPORT_RUNS]:
        run_cells.append(
            json.dumps(
                {
                    "run_index": run.run_index,
                    "query": run.query,
                    "raw_response": run.raw_response,
                    "parsed_response": run.parsed_response,
                    "predicted_answer": run.predicted_answer,
                    "result": int(run.is_correct),
                    "error": run.error,
                    "latency_seconds": run.latency_seconds,
                },
                ensure_ascii=False,
            )
        )

    while len(run_cells) < MAX_EXPORT_RUNS:
        run_cells.append("")

    sheet.append(
        base_row
        + run_cells
        + [
            artifact.summary.correct_count,
            artifact.summary.run_count,
            artifact.summary.accuracy,
        ]
    )

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["model", "run_count", "correct_count", "accuracy", "generated_at"])
    summary_sheet.append(
        [
            artifact.summary.model_name,
            artifact.summary.run_count,
            artifact.summary.correct_count,
            artifact.summary.accuracy,
            artifact.summary.generated_at.isoformat(),
        ]
    )

    validation_sheet = workbook.create_sheet("Validation")
    validation_sheet.append(["rule", "status", "detail"])
    for result in artifact.validation_results:
        validation_sheet.append([result["rule"], result["status"], result["detail"]])

    workbook.save(Path(output_path))
